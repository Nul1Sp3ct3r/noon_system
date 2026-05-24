"""
reports.py — standalone reporting utility module for Noon Financial.
Imports only from database.py (no circular imports).
"""

import io
import os
from datetime import datetime
from database import get_db

# ---------------------------------------------------------------------------
# Arabic month names
# ---------------------------------------------------------------------------

ARABIC_MONTHS = {
    '01': 'يناير', '02': 'فبراير', '03': 'مارس',
    '04': 'أبريل', '05': 'مايو',   '06': 'يونيو',
    '07': 'يوليو', '08': 'أغسطس', '09': 'سبتمبر',
    '10': 'أكتوبر','11': 'نوفمبر','12': 'ديسمبر',
}


def format_month_ar(month_str):
    """Convert '2026-04' → 'أبريل 2026'."""
    if not month_str or len(month_str) < 7:
        return month_str or ''
    year, mon = month_str[:4], month_str[5:7]
    return f"{ARABIC_MONTHS.get(mon, mon)} {year}"


# ---------------------------------------------------------------------------
# Shared filter helpers
# ---------------------------------------------------------------------------

def _date_where(from_date, to_date, col, prefix_len=10):
    """Return (conditions_list, params_list) for a date range filter."""
    conditions, params = [], []
    if from_date:
        conditions.append(f"SUBSTR({col},1,{prefix_len}) >= ?")
        params.append(from_date)
    if to_date:
        conditions.append(f"SUBSTR({col},1,{prefix_len}) <= ?")
        params.append(to_date)
    return conditions, params


def _brand_where(brand, alias='p'):
    """Return (conditions_list, params_list) for a brand filter."""
    conditions, params = [], []
    if brand:
        conditions.append(f"{alias}.brand_en = ?")
        params.append(brand)
    return conditions, params


# ---------------------------------------------------------------------------
# Data functions
# ---------------------------------------------------------------------------

def get_vat_data(db_path, from_date='', to_date=''):
    """Monthly VAT breakdown using verified formulas.

    output_vat      = SUM(net_proceeds WHERE delivered) * 15/115
    input_vat_noon  = SUM(ABS(referral_fee)+ABS(fbn_outbound_fee)) * 0.15
    input_vat_supp  = SUM(invoices.vat_amount) for same month
    net_vat         = output_vat - input_vat_noon - input_vat_supp
    """
    db = get_db(db_path)

    # Orders query
    cond, params = _date_where(from_date, to_date, 'o.ordered_date', 10)
    where = ("WHERE " + " AND ".join(cond) + " AND o.ordered_date != ''") if cond else "WHERE o.ordered_date != ''"

    orders_sql = f"""
        SELECT
            strftime('%Y-%m', o.ordered_date) AS month,
            COALESCE(SUM(CASE WHEN o.item_status='delivered'
                              THEN o.net_proceeds ELSE 0 END), 0.0) AS sales_incl,
            COALESCE(SUM(ABS(o.referral_fee) + ABS(o.fbn_outbound_fee)), 0.0) AS fees_excl
        FROM orders o
        {where}
        GROUP BY strftime('%Y-%m', o.ordered_date)
        ORDER BY month
    """
    order_rows = db.execute(orders_sql, params).fetchall()

    # Supplier VAT from invoices (grouped by month of invoice_date)
    inv_cond, inv_params = _date_where(from_date, to_date, 'i.invoice_date', 10)
    where_inv = ("WHERE " + " AND ".join(inv_cond)) if inv_cond else ""
    inv_sql = f"""
        SELECT
            strftime('%Y-%m', i.invoice_date) AS month,
            COALESCE(SUM(i.vat_amount), 0.0) AS supplier_vat
        FROM invoices i
        {where_inv}
        GROUP BY strftime('%Y-%m', i.invoice_date)
    """
    inv_rows = db.execute(inv_sql, inv_params).fetchall()

    # Statement-level fee VAT from monthly-format imports
    nsf_cond2, nsf_params2 = _date_where(from_date, to_date, 'statement_date', 7)
    nsf_where2 = (
        "WHERE " + " AND ".join(nsf_cond2) + " AND statement_date != ''"
    ) if nsf_cond2 else "WHERE statement_date != ''"
    nsf_vat_sql = f"""
        SELECT
            strftime('%Y-%m', statement_date) AS month,
            COALESCE(SUM(ABS(excl_vat)),   0.0) AS fees_excl_stmt,
            COALESCE(SUM(ABS(vat_amount)), 0.0) AS stmt_vat
        FROM noon_statement_fees
        {nsf_where2}
        GROUP BY strftime('%Y-%m', statement_date)
    """
    nsf_vat_rows = db.execute(nsf_vat_sql, nsf_params2).fetchall()
    db.close()

    supp_vat_by_month  = {r['month']: float(r['supplier_vat']) for r in inv_rows}
    stmt_fees_by_month = {
        r['month']: {'fees_excl': float(r['fees_excl_stmt']), 'vat': float(r['stmt_vat'])}
        for r in nsf_vat_rows
    }

    result = []
    for r in order_rows:
        month      = r['month']
        sales_incl = float(r['sales_incl'])
        fees_excl  = float(r['fees_excl'])

        stmt = stmt_fees_by_month.get(month, {'fees_excl': 0.0, 'vat': 0.0})

        output_vat     = round(sales_incl * 15 / 115, 2)
        input_vat_noon = round(fees_excl * 0.15 + stmt['vat'], 2)
        input_vat_supp = round(supp_vat_by_month.get(month, 0.0), 2)
        net_vat        = round(output_vat - input_vat_noon - input_vat_supp, 2)

        result.append({
            'month':          month,
            'month_ar':       format_month_ar(month),
            'sales_incl':     round(sales_incl, 2),
            'output_vat':     output_vat,
            'fees_excl':      round(fees_excl + stmt['fees_excl'], 2),
            'input_vat_noon': input_vat_noon,
            'input_vat_supp': input_vat_supp,
            'net_vat':        net_vat,
            'status':         'payable' if net_vat > 0 else 'refundable',
        })
    return result


def get_settlements_data(db_path, from_date='', to_date=''):
    """Per-import-batch settlement reconciliation.

    Query drives from orders (has historical data) and LEFT JOINs imported_files
    for metadata — imported_files may be empty for historical imports.

    Verified formula:
      calc_net = gross_sales - noon_fees_excl_vat
    VAT on fees is informational only; it is deducted at statement level and is
    NOT included in per-row total_payment, so must not be subtracted here.
    """
    db = get_db(db_path)

    # Filter on import_batch (timestamp), prefix 7 for YYYY-MM month inputs
    cond, params = _date_where(from_date, to_date, 'o.import_batch', 7)
    where = ("WHERE " + " AND ".join(cond)) if cond else ""

    sql = f"""
        SELECT
            o.import_batch,
            f.id            AS file_id,
            f.statement_nr,
            f.statement_date,
            f.filename,
            f.rows_added,
            COUNT(o.id)     AS order_rows,
            COALESCE(SUM(CASE WHEN o.item_status='delivered'
                              THEN o.net_proceeds ELSE 0 END), 0.0) AS gross_sales,
            COALESCE(SUM(ABS(o.referral_fee)),      0.0) AS referral_fees,
            COALESCE(SUM(ABS(o.fbn_outbound_fee)),  0.0) AS fbn_fees,
            COALESCE(SUM(o.total_payment),           0.0) AS actual_payout
        FROM orders o
        LEFT JOIN imported_files f ON f.imported_at = o.import_batch
        {where}
        GROUP BY o.import_batch
        ORDER BY o.import_batch DESC
    """
    rows = db.execute(sql, params).fetchall()

    # Statement-level fees per import_batch (monthly-format imports)
    nsf_batch_rows = db.execute("""
        SELECT import_batch,
               COALESCE(SUM(ABS(excl_vat)),   0.0) AS fees_excl,
               COALESCE(SUM(ABS(vat_amount)), 0.0) AS fee_vat,
               COALESCE(SUM(ABS(incl_vat)),   0.0) AS fees_incl
        FROM noon_statement_fees
        GROUP BY import_batch
    """).fetchall()
    db.close()

    nsf_by_batch = {
        r['import_batch']: {
            'fees_excl': float(r['fees_excl']),
            'vat':       float(r['fee_vat']),
            'fees_incl': float(r['fees_incl']),
        }
        for r in nsf_batch_rows
    }

    result = []
    for r in rows:
        gross_sales   = round(float(r['gross_sales']),   2)
        referral_fees = round(float(r['referral_fees']), 2)
        fbn_fees      = round(float(r['fbn_fees']),      2)
        actual_payout = round(float(r['actual_payout']), 2)

        stmt = nsf_by_batch.get(r['import_batch'], {'fees_excl': 0, 'vat': 0, 'fees_incl': 0})
        is_monthly_batch = stmt['fees_excl'] > 0

        if is_monthly_batch:
            total_fees  = round(stmt['fees_excl'], 2)
            vat_on_fees = round(stmt['vat'], 2)
        else:
            total_fees  = round(referral_fees + fbn_fees, 2)
            vat_on_fees = round(total_fees * 0.15, 2)

        our_net  = round(gross_sales - total_fees, 2)
        mismatch = round(abs(our_net - actual_payout), 2) if not is_monthly_batch else 0

        result.append({
            'import_batch':    r['import_batch'],
            'statement_nr':    r['statement_nr'] or '—',
            'statement_date':  r['statement_date'] or '—',
            'filename':        r['filename'] or r['import_batch'],
            'rows_added':      r['rows_added'] or r['order_rows'],
            'gross_sales':     gross_sales,
            'referral_fees':   referral_fees,
            'fbn_fees':        fbn_fees,
            'stmt_fees':       round(stmt['fees_excl'], 2),
            'total_fees':      total_fees,
            'vat_on_fees':     vat_on_fees,
            'our_net':         our_net,
            'actual_payout':   actual_payout,
            'mismatch':        mismatch,
            'has_mismatch':    mismatch > 1.0,
            'is_monthly_batch': is_monthly_batch,
        })
    return result


def get_profitability_data(db_path, from_date='', to_date='',
                           sku_search='', badge_filter=''):
    """Per-SKU profitability with date filtering and new badge thresholds.

    Reuses compute_product_metrics via an inline import to avoid circular deps.
    Date filter is applied in the JOIN condition so all products still appear.
    Badge thresholds (overrides compute_product_metrics defaults):
      profitable  = net_profit >= 2.0 SAR
      low_margin  = 0 <= net_profit < 2.0 SAR
      loss        = net_profit < 0
      missing_cost = no unit_cost or extra_costs set
    """
    from app import compute_product_metrics, VAT_RATE, VAT_FACTOR  # app imports db, not circular

    db = get_db(db_path)

    # Date conditions in the JOIN so all products appear even with no matching orders
    fd = from_date or ''
    td = to_date   or ''

    sql = """
        SELECT
            p.sku, p.partner_sku, p.brand_en, p.brand_ar,
            p.name_en, p.name_ar, p.unit_cost, p.extra_costs, p.notes,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN 1 ELSE 0 END), 0) AS units_sold,
            COALESCE(SUM(CASE WHEN o.item_status='returned'  THEN 1 ELSE 0 END), 0) AS units_returned,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN o.net_proceeds ELSE 0 END), 0.0) AS revenue,
            COALESCE(SUM(ABS(o.referral_fee) + ABS(o.fbn_outbound_fee)), 0.0) AS noon_fees
        FROM products p
        LEFT JOIN orders o
            ON  o.sku = p.sku
            AND ('' = ? OR SUBSTR(o.ordered_date, 1, 7) >= ?)
            AND ('' = ? OR SUBSTR(o.ordered_date, 1, 7) <= ?)
        GROUP BY p.sku
        ORDER BY p.name_en
    """
    rows = db.execute(sql, (fd, fd, td, td)).fetchall()

    # Check whether any noon_statement_fees exist in the filtered period
    nsf_where_p = "WHERE statement_date != ''"
    nsf_params_p = []
    if fd:
        nsf_where_p += " AND SUBSTR(statement_date, 1, 7) >= ?"
        nsf_params_p.append(fd)
    if td:
        nsf_where_p += " AND SUBSTR(statement_date, 1, 7) <= ?"
        nsf_params_p.append(td)
    nsf_cnt = db.execute(
        f"SELECT COUNT(*) FROM noon_statement_fees {nsf_where_p}", nsf_params_p
    ).fetchone()
    has_stmt_fees = int(nsf_cnt[0] if nsf_cnt else 0) > 0
    db.close()

    result = []
    for row in rows:
        m = compute_product_metrics(row)

        m['has_unallocated_fees'] = has_stmt_fees and m['noon_fees'] == 0

        # Apply profitability-specific badge thresholds
        if not m['has_cost']:
            m['badge'] = 'missing_cost'
        elif m['has_unallocated_fees']:
            m['badge'] = 'no_fees_allocated'
        elif m['net_profit'] >= 2.0:
            m['badge'] = 'profitable'
        elif m['net_profit'] >= 0:
            m['badge'] = 'low_margin'
        else:
            m['badge'] = 'loss'

        result.append(m)

    # Python-side filters (fast, data is small)
    if sku_search:
        q = sku_search.lower()
        result = [m for m in result
                  if q in m['sku'].lower()
                  or q in (m['name_en'] or '').lower()
                  or q in (m['name_ar'] or '').lower()]
    if badge_filter:
        result = [m for m in result if m['badge'] == badge_filter]

    return result


def get_pl_data(db_path, from_date='', to_date=''):
    """P&L grouped by month."""
    db = get_db(db_path)
    cond, params = _date_where(from_date, to_date, 'o.ordered_date', 10)
    where = ("WHERE " + " AND ".join(cond) + " AND o.ordered_date != ''") if cond else "WHERE o.ordered_date != ''"

    sql = f"""
        SELECT
            strftime('%Y-%m', o.ordered_date) AS month,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN o.net_proceeds ELSE 0 END), 0.0) AS revenue,
            COALESCE(SUM(ABS(o.referral_fee) + ABS(o.fbn_outbound_fee)), 0.0) AS fees,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN COALESCE(p.unit_cost, 0) ELSE 0 END), 0.0) AS cogs,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN COALESCE(p.extra_costs, 0) ELSE 0 END), 0.0) AS extra
        FROM orders o
        LEFT JOIN products p ON p.sku = o.sku
        {where}
        GROUP BY strftime('%Y-%m', o.ordered_date)
        ORDER BY month
    """
    rows = db.execute(sql, params).fetchall()
    db.close()

    result = []
    for r in rows:
        revenue = float(r['revenue'])
        fees = float(r['fees'])
        cogs = float(r['cogs'])
        extra = float(r['extra'])
        gross_profit = revenue - fees
        net_profit = gross_profit - cogs - extra
        margin_pct = round(net_profit / revenue * 100, 2) if revenue else None
        result.append({
            'month': r['month'],
            'month_ar': format_month_ar(r['month']),
            'revenue': round(revenue, 2),
            'fees': round(fees, 2),
            'cogs': round(cogs, 2),
            'extra': round(extra, 2),
            'gross_profit': round(gross_profit, 2),
            'net_profit': round(net_profit, 2),
            'margin_pct': margin_pct,
        })
    return result


def get_sales_data(db_path, from_date='', to_date='', brand='', sort_by='profit', status=''):
    """Sales breakdown per SKU."""
    db = get_db(db_path)

    date_cond, date_params = _date_where(from_date, to_date, 'o.ordered_date', 10)
    brand_cond, brand_params = _brand_where(brand, 'p')

    all_cond = date_cond + brand_cond
    if date_cond or brand_cond:
        where = "WHERE " + " AND ".join(all_cond)
    else:
        where = ""

    sql = f"""
        SELECT
            p.sku,
            p.partner_sku,
            p.name_en,
            p.name_ar,
            p.brand_en,
            p.brand_ar,
            p.unit_cost,
            p.extra_costs,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN 1 ELSE 0 END), 0) AS units_sold,
            COALESCE(SUM(CASE WHEN o.item_status='returned'  THEN 1 ELSE 0 END), 0) AS units_returned,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN o.net_proceeds ELSE 0 END), 0.0) AS revenue,
            COALESCE(SUM(ABS(o.referral_fee) + ABS(o.fbn_outbound_fee)), 0.0) AS noon_fees
        FROM products p
        LEFT JOIN orders o ON o.sku = p.sku
        {where}
        GROUP BY p.sku
    """
    rows = db.execute(sql, date_params + brand_params).fetchall()
    db.close()

    result = []
    for r in rows:
        units_sold = int(r['units_sold'] or 0)
        units_returned = int(r['units_returned'] or 0)
        revenue = float(r['revenue'] or 0)
        noon_fees = float(r['noon_fees'] or 0)
        unit_cost = float(r['unit_cost'] or 0)
        extra_costs = float(r['extra_costs'] or 0)
        cogs = unit_cost * units_sold
        net_profit = revenue - noon_fees - cogs - extra_costs
        margin_pct = round(net_profit / revenue * 100, 2) if revenue else None

        total_units = units_sold + units_returned
        return_rate = round(units_returned / total_units * 100, 2) if total_units else 0.0

        has_cost = unit_cost > 0 or extra_costs > 0
        if not has_cost:
            badge = 'unknown'
        elif net_profit > 0:
            badge = 'profitable'
        else:
            badge = 'loss'

        result.append({
            'sku': r['sku'],
            'partner_sku': r['partner_sku'] or '',
            'name_en': r['name_en'] or '',
            'name_ar': r['name_ar'] or '',
            'brand_en': r['brand_en'] or '',
            'brand_ar': r['brand_ar'] or '',
            'units_sold': units_sold,
            'units_returned': units_returned,
            'return_rate': return_rate,
            'revenue': round(revenue, 2),
            'noon_fees': round(noon_fees, 2),
            'cogs': round(cogs, 2),
            'net_profit': round(net_profit, 2),
            'margin_pct': margin_pct,
            'badge': badge,
        })

    # Filter by status badge
    if status:
        result = [r for r in result if r['badge'] == status]

    # Sort
    sort_key_map = {
        'profit': lambda x: x['net_profit'],
        'revenue': lambda x: x['revenue'],
        'units': lambda x: x['units_sold'],
    }
    key_fn = sort_key_map.get(sort_by, sort_key_map['profit'])
    result.sort(key=key_fn, reverse=True)

    return result


def get_fees_data(db_path, from_date='', to_date='', brand=''):
    """Fee breakdown per SKU."""
    db = get_db(db_path)

    date_cond, date_params = _date_where(from_date, to_date, 'o.ordered_date', 10)
    brand_cond, brand_params = _brand_where(brand, 'p')

    all_cond = date_cond + brand_cond
    where = ("WHERE " + " AND ".join(all_cond)) if all_cond else ""

    sql = f"""
        SELECT
            p.sku,
            p.partner_sku,
            p.name_en,
            p.name_ar,
            p.brand_en,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN 1 ELSE 0 END), 0) AS units_del,
            COALESCE(SUM(CASE WHEN o.item_status='returned'  THEN 1 ELSE 0 END), 0) AS units_ret,
            COALESCE(SUM(ABS(o.referral_fee)), 0.0) AS referral,
            COALESCE(SUM(ABS(o.fbn_outbound_fee)), 0.0) AS fbn,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN o.net_proceeds ELSE 0 END), 0.0) AS revenue
        FROM products p
        LEFT JOIN orders o ON o.sku = p.sku
        {where}
        GROUP BY p.sku
    """
    rows = db.execute(sql, date_params + brand_params).fetchall()
    db.close()

    result = []
    for r in rows:
        units_del = int(r['units_del'] or 0)
        units_ret = int(r['units_ret'] or 0)
        referral = float(r['referral'] or 0)
        fbn = float(r['fbn'] or 0)
        revenue = float(r['revenue'] or 0)
        total_fees = referral + fbn
        fee_pct = round(total_fees / revenue * 100, 2) if revenue else 0.0
        net_after_fees = round(revenue - total_fees, 2)

        result.append({
            'sku': r['sku'],
            'partner_sku': r['partner_sku'] or '',
            'name_en': r['name_en'] or '',
            'name_ar': r['name_ar'] or '',
            'brand_en': r['brand_en'] or '',
            'units_del': units_del,
            'units_ret': units_ret,
            'referral': round(referral, 2),
            'fbn': round(fbn, 2),
            'total_fees': round(total_fees, 2),
            'revenue': round(revenue, 2),
            'fee_pct': fee_pct,
            'net_after_fees': net_after_fees,
        })
    return result


def get_inventory_data(db_path, brand='', cost_min=None, cost_max=None):
    """Inventory & cost data per SKU."""
    db = get_db(db_path)

    brand_cond, brand_params = _brand_where(brand, 'p')
    cond = list(brand_cond)
    params = list(brand_params)

    if cost_min is not None:
        try:
            cond.append("p.unit_cost >= ?")
            params.append(float(cost_min))
        except (TypeError, ValueError):
            pass
    if cost_max is not None:
        try:
            cond.append("p.unit_cost <= ?")
            params.append(float(cost_max))
        except (TypeError, ValueError):
            pass

    where = ("WHERE " + " AND ".join(cond)) if cond else ""

    sql = f"""
        SELECT
            p.sku,
            p.partner_sku,
            p.name_en,
            p.name_ar,
            p.brand_en,
            p.unit_cost,
            p.extra_costs,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN 1 ELSE 0 END), 0) AS units_sold,
            COALESCE(SUM(CASE WHEN o.item_status='returned'  THEN 1 ELSE 0 END), 0) AS units_returned,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN COALESCE(p.unit_cost, 0) ELSE 0 END), 0.0) AS total_cogs,
            COALESCE(SUM(CASE WHEN o.item_status='delivered' THEN COALESCE(p.extra_costs, 0) ELSE 0 END), 0.0) AS total_extra
        FROM products p
        LEFT JOIN orders o ON o.sku = p.sku
        {where}
        GROUP BY p.sku
    """
    rows = db.execute(sql, params).fetchall()
    db.close()

    result = []
    for r in rows:
        unit_cost = float(r['unit_cost'] or 0)
        extra_costs = float(r['extra_costs'] or 0)
        total_per_unit = unit_cost + extra_costs
        units_sold = int(r['units_sold'] or 0)
        units_returned = int(r['units_returned'] or 0)
        total_cogs = float(r['total_cogs'] or 0)
        total_extra = float(r['total_extra'] or 0)
        total_investment = total_cogs + total_extra

        result.append({
            'sku': r['sku'],
            'partner_sku': r['partner_sku'] or '',
            'name_en': r['name_en'] or '',
            'name_ar': r['name_ar'] or '',
            'brand_en': r['brand_en'] or '',
            'unit_cost': round(unit_cost, 2),
            'extra_costs': round(extra_costs, 2),
            'total_per_unit': round(total_per_unit, 2),
            'units_sold': units_sold,
            'units_returned': units_returned,
            'total_cogs': round(total_cogs, 2),
            'total_extra': round(total_extra, 2),
            'total_investment': round(total_investment, 2),
        })
    return result


def get_invoices_report_data(db_path, supplier='', from_date='', to_date=''):
    """Invoice list with per-invoice aggregates."""
    db = get_db(db_path)

    cond, params = [], []
    if supplier:
        cond.append("i.supplier_name = ?")
        params.append(supplier)
    date_cond, date_params = _date_where(from_date, to_date, 'i.invoice_date', 10)
    cond += date_cond
    params += date_params

    where = ("WHERE " + " AND ".join(cond)) if cond else ""

    sql = f"""
        SELECT
            i.id,
            i.invoice_nr,
            i.supplier_name,
            i.invoice_date,
            i.total_amount,
            i.vat_amount,
            i.pdf_filename,
            COUNT(ii.id) AS item_count
        FROM invoices i
        LEFT JOIN invoice_items ii ON ii.invoice_id = i.id
        {where}
        GROUP BY i.id
        ORDER BY i.supplier_name, i.invoice_date
    """
    rows = db.execute(sql, params).fetchall()
    db.close()

    result = []
    for r in rows:
        total_amount = float(r['total_amount'] or 0)
        vat_amount = float(r['vat_amount'] or 0)
        net_amount = round(total_amount - vat_amount, 2)
        result.append({
            'id': r['id'],
            'invoice_nr': r['invoice_nr'] or '',
            'supplier_name': r['supplier_name'] or '',
            'invoice_date': r['invoice_date'] or '',
            'total_amount': round(total_amount, 2),
            'vat_amount': round(vat_amount, 2),
            'net_amount': net_amount,
            'item_count': int(r['item_count'] or 0),
            'has_pdf': bool(r['pdf_filename']),
        })
    return result


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_excel(sheet_name, headers, rows, totals_row=None,
                currency_cols=None, pct_cols=None, filename=None):
    """
    Build a styled Excel workbook and return (BytesIO, filename).
    currency_cols and pct_cols are 1-based column index lists.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    currency_cols = currency_cols or []
    pct_cols = pct_cols or []

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_fill = PatternFill('solid', fgColor='1A3A5C')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    alt_fill = PatternFill('solid', fgColor='F8F9FA')
    totals_fill = PatternFill('solid', fgColor='FFF3CD')
    totals_font = Font(bold=True, name='Arial', size=10)
    normal_font = Font(name='Arial', size=9)
    thin_side = Side(style='thin', color='DEE2E6')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Header row
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        ws.append(list(row))
        row_fill = None if row_idx % 2 == 0 else alt_fill
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            if row_fill:
                cell.fill = row_fill
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in currency_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif col_idx in pct_cols:
                cell.number_format = '0.00'
                cell.alignment = right_align
            else:
                cell.alignment = left_align

    # Totals row
    if totals_row:
        ws.append(list(totals_row))
        totals_row_idx = ws.max_row
        for col_idx, cell in enumerate(ws[totals_row_idx], start=1):
            cell.fill = totals_fill
            cell.font = totals_font
            cell.border = thin_border
            if col_idx in currency_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif col_idx in pct_cols:
                cell.number_format = '0.00'
                cell.alignment = right_align
            else:
                cell.alignment = left_align

    # Auto-fit column widths (capped at 45)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = min(max_len + 4, 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Freeze header row
    ws.freeze_panes = 'A2'

    if not filename:
        today = datetime.now().strftime('%Y%m%d')
        filename = f'noon_{sheet_name.lower().replace(" ", "_")}_{today}.xlsx'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename


# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------

# Module-level flags to avoid repeated font registration
_pdf_font_registered = False
_pdf_font_name = 'Helvetica'
_arabic_reshape_available = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _reportlab_available = True
except ImportError:
    _reportlab_available = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    _arabic_reshape_available = True
except ImportError:
    _arabic_reshape_available = False


def _ar(text):
    """Reshape Arabic text for correct RTL rendering in PDF."""
    text = str(text) if text is not None else ''
    if not text:
        return text
    if _arabic_reshape_available:
        try:
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        except Exception:
            return text
    return text


def _ensure_font_registered():
    """Register an Arabic-capable TTF font once, return the font name to use."""
    global _pdf_font_registered, _pdf_font_name

    if _pdf_font_registered:
        return _pdf_font_name

    candidate_paths = [
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\tahoma.ttf',
        '/mnt/c/Windows/Fonts/arial.ttf',
        '/mnt/c/Windows/Fonts/tahoma.ttf',
    ]

    # Also search common Linux font dirs
    for font_dir in ['/usr/share/fonts', '/usr/local/share/fonts', os.path.expanduser('~/.fonts')]:
        if os.path.isdir(font_dir):
            for root, dirs, files in os.walk(font_dir):
                for f in files:
                    if f.lower().endswith('.ttf'):
                        candidate_paths.append(os.path.join(root, f))
                break  # only top level to keep it fast

    registered = False
    for path in candidate_paths:
        if os.path.isfile(path):
            try:
                pdfmetrics.registerFont(TTFont('NoonArabic', path))
                _pdf_font_name = 'NoonArabic'
                registered = True
                break
            except Exception:
                continue

    if not registered:
        _pdf_font_name = 'Helvetica'

    _pdf_font_registered = True
    return _pdf_font_name


def build_pdf(title_ar, headers, rows, totals_row=None, summary=None):
    """
    Build a PDF report and return (BytesIO, None) or (None, error_str).
    """
    if not _reportlab_available:
        return None, "reportlab not installed"

    try:
        font_name = _ensure_font_registered()

        buf = io.BytesIO()
        page_size = landscape(A4)
        margin = 15 * mm

        def _make_header_footer(canvas_obj, doc_obj):
            canvas_obj.saveState()
            width, height = page_size
            # Header
            canvas_obj.setFont(font_name, 10)
            canvas_obj.setFillColor(colors.HexColor('#1A3A5C'))
            canvas_obj.drawCentredString(width / 2, height - 10 * mm, _ar('نظام noon المالي'))
            canvas_obj.setFont(font_name, 12)
            canvas_obj.drawCentredString(width / 2, height - 18 * mm, _ar(title_ar))
            canvas_obj.setFont(font_name, 8)
            canvas_obj.setFillColor(colors.grey)
            date_str = datetime.now().strftime('%Y-%m-%d')
            canvas_obj.drawCentredString(width / 2, height - 24 * mm, date_str)
            # Footer
            canvas_obj.setFont(font_name, 8)
            canvas_obj.setFillColor(colors.grey)
            page_num = canvas_obj.getPageNumber()
            canvas_obj.drawCentredString(width / 2, 8 * mm, f'{page_num}')
            canvas_obj.restoreState()

        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            leftMargin=margin, rightMargin=margin,
            topMargin=35 * mm, bottomMargin=20 * mm,
        )

        story = []

        # Summary table if provided
        if summary:
            summary_data = [[_ar(k), _ar(str(v))] for k, v in summary.items()]
            summary_table = Table(summary_data, colWidths=[80 * mm, 80 * mm])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 5 * mm))

        # Build table data: header + rows + optional totals
        ar_headers = [_ar(h) for h in headers]
        table_data = [ar_headers]
        for row in rows:
            table_data.append([_ar(str(cell) if cell is not None else '—') for cell in row])
        if totals_row:
            table_data.append([_ar(str(cell) if cell is not None else '') for cell in totals_row])

        # Compute column widths
        page_width = page_size[0] - 2 * margin
        num_cols = len(headers)
        col_width = page_width / num_cols if num_cols else page_width
        col_widths = [col_width] * num_cols

        main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

        table_style_cmds = [
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A3A5C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
            ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

        # Totals row styling
        if totals_row:
            last_row_idx = len(table_data) - 1
            table_style_cmds += [
                ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor('#FFF3CD')),
                ('FONTNAME', (0, last_row_idx), (-1, last_row_idx), font_name),
                ('FONTSIZE', (0, last_row_idx), (-1, last_row_idx), 8),
            ]

        main_table.setStyle(TableStyle(table_style_cmds))
        story.append(main_table)

        doc.build(story, onFirstPage=_make_header_footer, onLaterPages=_make_header_footer)
        buf.seek(0)
        return buf, None

    except Exception as e:
        return None, str(e)
